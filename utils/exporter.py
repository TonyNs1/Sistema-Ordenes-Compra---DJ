import os
import json
import pandas as pd
import platform
from openpyxl import load_workbook
from openpyxl.styles import Alignment
import xlwings as xw

def exportar_usando_machote(df: pd.DataFrame, nombre_archivo: str, carpeta_destino: str, ruta_machote: str) -> str:
    try:
        # Validar columnas requeridas
        requeridas = ["Código", "Cantidad comprada", "Costo unitario de la compra", "Descuento"]
        for col in requeridas:
            if col not in df.columns:
                raise ValueError(f"❌ Faltan columnas requeridas en el DataFrame: {col}")

        df = df.fillna(0)

        # Ruta final
        ruta_salida = os.path.join(carpeta_destino, nombre_archivo)

        # Cargar machote
        wb = load_workbook(ruta_machote)
        ws = wb.active

        fila = 2
        for _, row in df.iterrows():
            codigo_str = str(row["Código"]).replace(".0", "") if isinstance(row["Código"], (int, float)) else str(row["Código"])
            ws.cell(row=fila, column=1, value=codigo_str)
            ws.cell(row=fila, column=2, value=round(float(row["Cantidad comprada"]), 2))
            ws.cell(row=fila, column=3, value=round(float(row["Costo unitario de la compra"]), 2))
            ws.cell(row=fila, column=4, value=round(float(row["Descuento"]), 2))
            fila += 1

        for row in ws.iter_rows(min_row=2, max_row=fila - 1, min_col=1, max_col=4):
            for cell in row:
                cell.alignment = Alignment(horizontal="center", vertical="center")

        wb.save(ruta_salida)
        wb.close()

        # Solo se aplica si estás en Windows
        ajustar_excel_con_excel(ruta_salida)

        return ruta_salida

    except Exception as e:
        raise RuntimeError(f"❌ Error al exportar usando la plantilla. Detalles: {e}")

def ajustar_excel_con_excel(ruta: str):
    if platform.system() != "Windows":
        return
    try:
        app = xw.App(visible=False)
        libro = app.books.open(ruta)
        libro.save()
        libro.close()
        app.quit()
    except Exception as e:
        print(f"⚠️ Error al ajustar con Excel: {e}")

def cargar_historial_ordenes(ruta_historial: str) -> list:
    if not os.path.exists(ruta_historial):
        return []
    with open(ruta_historial, "r", encoding="utf-8") as f:
        return json.load(f)

def exportar_orden_excel(df: pd.DataFrame, ruta: str) -> None:
    df.to_excel(ruta, index=False)

